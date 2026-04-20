"""
PDFWala V10.0
utils/office_utils.py — Excel / Office document helpers.
"""

from datetime import datetime, date
from decimal import Decimal
from typing import Any, List, Optional
import hashlib


def coerce_cell_value(value: Any) -> Any:
    """
    Coerce an Excel cell value to a JSON-serialisable Python type.
    Formerly _coerce_cell_value().
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value == int(value) and abs(value) < 1e15:
            return int(value)
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        f = float(value)
        return int(f) if f == int(f) else f
    return str(value)


def coerce_cell_for_csv(value: Any) -> str:
    """
    Coerce an Excel cell value to a CSV-safe string.
    Formerly _coerce_cell_for_csv().
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value == int(value) and abs(value) < 1e15:
            return str(int(value))
        return f"{value:.10g}"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def detect_excel_headers(rows: List[tuple]) -> Optional[List[str]]:
    """
    Heuristically detect whether the first row of an Excel sheet is a header.
    Returns a normalised list of header names, or None if no header detected.
    """
    if not rows or not rows[0]:
        return None
    first = rows[0]
    first_all_str = all(v is None or isinstance(v, str) for v in first)
    if not first_all_str:
        return None
    seen: dict = {}
    headers = []
    for i, h in enumerate(first):
        base = str(h).strip() if h is not None else ""
        if not base:
            base = f"col_{i}"
        cnt = seen.get(base, 0)
        seen[base] = cnt + 1
        headers.append(base if cnt == 0 else f"{base}_{cnt}")
    return headers


def is_excel_formula(value: Any) -> bool:
    """Return True if the cell value looks like an Excel formula."""
    return isinstance(value, str) and value.startswith("=")


def clean_excel_data(rows: List[tuple]) -> List[tuple]:
    """
    Remove fully empty rows and trailing empty columns from a row-list.
    """
    if not rows:
        return []
    filtered = [r for r in rows if any(v is not None and v != "" for v in r)]
    if not filtered:
        return []
    max_col = 0
    for row in filtered:
        for i, v in enumerate(row):
            if v is not None and v != "":
                max_col = max(max_col, i)
    return [row[: max_col + 1] for row in filtered]


def prepare_excel_for_pdf(input_path: str, output_path: str = None) -> str:
    """
    Pre-process Excel file for optimal PDF conversion.
    - Auto-fits all columns
    - Sets print area to used range
    - Scales to fit page width
    - Sets landscape orientation
    
    Args:
        input_path: Path to the original Excel file
        output_path: Optional output path for the prepared file
    
    Returns:
        Path to the prepared Excel file
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
    except ImportError:
        return input_path

    if output_path is None:
        output_path = input_path.replace('.xlsx', '_prepared.xlsx').replace('.xls', '_prepared.xlsx')

    wb = load_workbook(input_path)
    
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            try:
                col_letter = get_column_letter(col[0].column)
            except (IndexError, AttributeError):
                continue
            
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = min(cell_length, 50)
            
            adjusted_width = max_length + 2
            if adjusted_width > 0:
                sheet.column_dimensions[col_letter].width = adjusted_width

        if sheet.max_row > 0 and sheet.max_column > 0:
            print_area = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            sheet.print_area = print_area

        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        
        sheet.page_margins = PageMargins(
            left=0.25, right=0.25,
            top=0.5, bottom=0.5,
            header=0.3, footer=0.3
        )

    wb.save(output_path)
    wb.close()
    
    return output_path


# ============================================================================
# PDF TO EXCEL HELPER FUNCTIONS (10/10 UPGRADE)
# ============================================================================

def _is_structured_table(table: list) -> bool:
    """Smart table detection - checks row consistency and header presence."""
    if not table or len(table) < 2:
        return False
    
    clean_table = [_clean_row(row) for row in table if not _is_noise_row(row)]
    if len(clean_table) < 2:
        return False
    
    col_lengths = [len(row) for row in clean_table if row]
    if not col_lengths:
        return False
    
    if max(col_lengths) - min(col_lengths) > 1:
        return False
    
    header = clean_table[0]
    if not any(str(c).strip().isalpha() for c in header if c):
        return False
    
    return True


def _clean_row(row: list) -> list:
    """
    Clean row values WITHOUT removing columns.
    Preserves column alignment by keeping empty cells as empty strings.
    """
    if not row:
        return []
    return [str(c).strip() if c is not None and str(c).strip() else "" for c in row]


def _is_noise_row(row: list) -> bool:
    """Check if row is just empty cells or noise."""
    if not row:
        return True
    return not any(str(c).strip() for c in row if c is not None)


def _smart_cast(value: any) -> any:
    """Convert string numbers to actual int/float for Excel."""
    if value is None:
        return ""
    
    str_val = str(value).strip()
    if not str_val:
        return ""
    
    try:
        if '.' not in str_val:
            return int(str_val)
    except ValueError:
        pass
    
    try:
        return float(str_val)
    except ValueError:
        pass
    
    return str_val


def _normalize_header(header: list) -> list:
    """Clean and standardize header names."""
    normalized = []
    seen = {}
    
    for col in header:
        name = str(col).strip().title() if col else "Column"
        if not name:
            name = "Column"
        
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        
        normalized.append(name)
    
    return normalized


def _get_table_signature(table: list) -> str:
    """
    Generate unique signature for table deduplication.
    Uses first two rows + row count to avoid collisions.
    """
    if not table or len(table) < 2:
        return ""
    
    sig_rows = table[:2]
    sig_str = str([[str(c)[:20] if c else "" for c in row] for row in sig_rows])
    sig_str += str(len(table))  # Add row count to avoid hash collisions
    return hashlib.md5(sig_str.encode()).hexdigest()


def _merge_tables(existing_table: list, new_table: list) -> list:
    """Merge continuation of multi-page table."""
    if not existing_table or not new_table:
        return existing_table or new_table
    
    existing_header = _normalize_header(existing_table[0])
    new_header = _normalize_header(new_table[0])
    
    if existing_header == new_header:
        return existing_table + new_table[1:]
    
    return existing_table + new_table


def _write_optimized_sheet(ws, table: list, method: str = "auto"):
    """Write table to worksheet with smart formatting and type detection."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    if not table:
        return
    
    header = _normalize_header(table[0])
    data_rows = table[1:] if len(table) > 1 else []
    
    for col_idx, col_name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_idx, row in enumerate(data_rows, 2):
        clean_row = _clean_row(row)
        for col_idx, value in enumerate(clean_row, 1):
            if col_idx > len(header):
                break
            cast_value = _smart_cast(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=cast_value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    for col in ws.iter_cols(max_row=min(ws.max_row, 200)):
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = min(cell_len, 50)
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width
    
    ws.freeze_panes = 'A2'
